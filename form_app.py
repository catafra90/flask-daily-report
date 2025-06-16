
from flask import Flask, render_template, request, redirect, url_for, send_file, session
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import pandas as pd
import requests

app = Flask(__name__)
app.secret_key = 'your_secure_random_key_here'

CHAT_WEBHOOK_URL = "https://chat.googleapis.com/v1/spaces/4ZsvACAAAAE/messages?key=AIzaSyDdI0hCZtE6vySjMm-WEfRq3CPzqKqqsHI&token=QAsMtcgO05jwCeg2HXcDrCK7ngmtq0vpQwguobG8-vU"

def send_to_google_chat(message):
    try:
        payload = {"text": message}
        headers = {'Content-Type': 'application/json'}
        response = requests.post(CHAT_WEBHOOK_URL, headers=headers, json=payload)
        print("✅ Sent to Google Chat:", response.status_code)
    except Exception as e:
        print("❌ Failed to send to Google Chat:", e)

@app.route('/')
def index():
    return render_template('index.html', active_page='home')

@app.route('/step1', methods=['GET', 'POST'])
def step1():
    if request.method == 'POST':
        session['sales'] = {
            'client_name': request.form.getlist('client_name'),
            'package': request.form.getlist('package'),
            'revenue': request.form.getlist('revenue')
        }
        return redirect(url_for('step2'))
    return render_template('step1.html', sales=session.get('sales', {}), active_page='daily-report')

@app.route('/step2', methods=['GET', 'POST'])
def step2():
    if request.method == 'POST':
        session['leads'] = {
            'name': request.form.getlist('lead_name'),
            'date': request.form.getlist('lead_date'),
            'source': request.form.getlist('lead_source')
        }
        return redirect(url_for('step3'))
    return render_template('step2.html', leads=session.get('leads', {}), active_page='daily-report')

@app.route('/step3', methods=['GET', 'POST'])
def step3():
    if request.method == 'POST':
        session['consultations'] = {
            'name': request.form.getlist('consultation_name'),
            'outcome': request.form.getlist('consultation_outcome'),
            'source': request.form.getlist('consultation_source')
        }
        return redirect(url_for('step4'))
    return render_template('step3.html', consultations=session.get('consultations', {}), active_page='daily-report')

@app.route('/step4', methods=['GET', 'POST'])
def step4():
    if request.method == 'POST':
        session['opportunities'] = {
            'name': request.form.getlist('opportunity_name'),
            'provider': request.form.getlist('opportunity_provider'),
            'description': request.form.getlist('opportunity_description')
        }
        return redirect(url_for('step5'))
    return render_template('step4.html', opportunities=session.get('opportunities', {}), active_page='daily-report')

@app.route('/step5', methods=['GET', 'POST'])
def step5():
    if request.method == 'POST':
        session['attendance_done'] = request.form.get('attendance_done', '')
        session['no_show'] = request.form.get('no_show', '')
        return redirect(url_for('submitted'))

    attendance_data = {
        'attendance_done': session.get('attendance_done', ''),
        'no_show': session.get('no_show', '')
    }

    return render_template('step5.html', attendance=attendance_data, active_page='daily-report')

@app.route('/submitted')
def submitted():
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        filepath = "reports.xlsx"

        if os.path.exists(filepath):
            wb = load_workbook(filepath)
        else:
            wb = Workbook()
            for sheet in wb.sheetnames:
                del wb[sheet]
            for name, headers in {
                "Sales": ["Date", "Client Name", "Package", "Revenue"],
                "Leads": ["Date", "Lead Name", "Lead Date", "Lead Source"],
                "Consultations": ["Date", "Consultation Name", "Consultation Outcome", "Consultation Source"],
                "Opportunities": ["Date", "Opportunity Name", "Opportunity Provider", "Opportunity Description"],
                "Attendance": ["Date", "Attendance Done", "No Show"]
            }.items():
                ws = wb.create_sheet(title=name)
                ws.append(headers)

        # --- SALES ---
        sales = session.get('sales', {})
        ws = wb["Sales"]
        for i in range(len(sales.get('client_name', []))):
            if sales['client_name'][i] or sales['package'][i] or sales['revenue'][i]:
                ws.append([timestamp, sales['client_name'][i], sales['package'][i], sales['revenue'][i]])

        # --- LEADS ---
        leads = session.get('leads', {})
        ws = wb["Leads"]
        for i in range(len(leads.get('name', []))):
            if leads['name'][i] or leads['date'][i] or leads['source'][i]:
                ws.append([timestamp, leads['name'][i], leads['date'][i], leads['source'][i]])

        # --- CONSULTATIONS ---
        consultations = session.get('consultations', {})
        ws = wb["Consultations"]
        for i in range(len(consultations.get('name', []))):
            if consultations['name'][i] or consultations['outcome'][i] or consultations['source'][i]:
                ws.append([timestamp, consultations['name'][i], consultations['outcome'][i], consultations['source'][i]])

        # --- OPPORTUNITIES ---
        opportunities = session.get('opportunities', {})
        ws = wb["Opportunities"]
        for i in range(len(opportunities.get('name', []))):
            if opportunities['name'][i] or opportunities['provider'][i] or opportunities['description'][i]:
                ws.append([timestamp, opportunities['name'][i], opportunities['provider'][i], opportunities['description'][i]])

        # --- ATTENDANCE ---
        attendance_done = session.get('attendance_done', '')
        no_show = session.get('no_show', '')
        if attendance_done or no_show:
            ws = wb["Attendance"]
            ws.append([timestamp, attendance_done, no_show])

        wb.save(filepath)
        print("✅ Excel report saved as reports.xlsx")

        # --- Google Chat Message ---
        message = f"**Daily Report Submitted**\n\n"
        if sales:
            message += "*Sales:*\n"
            for i in range(len(sales.get('client_name', []))):
                if sales['client_name'][i] or sales['package'][i] or sales['revenue'][i]:
                    message += f"- {sales['client_name'][i]} | {sales['package'][i]} | ${sales['revenue'][i]}\n"
        if leads:
            message += "\n*Leads:*\n"
            for i in range(len(leads.get('name', []))):
                if leads['name'][i] or leads['date'][i] or leads['source'][i]:
                    message += f"- {leads['name'][i]} | {leads['date'][i]} | {leads['source'][i]}\n"
        if consultations:
            message += "\n*Consultations:*\n"
            for i in range(len(consultations.get('name', []))):
                if consultations['name'][i] or consultations['outcome'][i] or consultations['source'][i]:
                    message += f"- {consultations['name'][i]} | {consultations['outcome'][i]} | {consultations['source'][i]}\n"
        if opportunities:
            message += "\n*Opportunities:*\n"
            for i in range(len(opportunities.get('name', []))):
                if opportunities['name'][i] or opportunities['provider'][i] or opportunities['description'][i]:
                    message += f"- {opportunities['name'][i]} | {opportunities['provider'][i]} | {opportunities['description'][i]}\n"
        if attendance_done or no_show:
            message += f"\n*Attendance:* {attendance_done} done, {no_show} no-show"

        send_to_google_chat(message)
        session.clear()

    except Exception as e:
        print("❌ Error in submission:", e)

    return render_template("submitted.html")

@app.route('/clients')
def clients():
    try:
        df = pd.read_excel("all_clients.xlsx")
        data = df.values.tolist()
        return render_template("clients.html", active_page='clients', clients=data)
    except Exception as e:
        print("❌ Error loading client list:", e)
        return render_template("clients.html", active_page='clients', clients=[])

if __name__ == '__main__':
    app.run(debug=True)
